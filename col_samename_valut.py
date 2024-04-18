# 这个程序用于将同一列中重复的名字，数值合并。

import os
from openpyxl import load_workbook


def process_excel_file(file_path):
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet_name == '地址组对象':
            # 存储已经出现的值
            seen_values = set()
            # 存储需要保留的行索引及对应的B列数值
            rows_to_keep = {}
            # 存储重复项及其数量
            duplicates = {}

            for row in sheet.iter_rows(min_row=1, max_row=35000, min_col=1, max_col=2, values_only=True):
                a_value, b_value = row
                if a_value in seen_values:
                    # 如果A列值已经出现过，则将B列值加到之前存储的行上
                    first_col_row = rows_to_keep[a_value]
                    first_col_row[1] = first_col_row[1] or 0  # Initialize to 0 if it's None
                    first_col_row[1] += b_value or 0  # Initialize to 0 if it's None
                    # 更新重复项的数量
                    duplicates[a_value] = duplicates.get(a_value, 0) + 1
                else:
                    # 如果A列值是第一次出现，则记录该行索引及B列值
                    seen_values.add(a_value)
                    rows_to_keep[a_value] = [a_value, b_value]

            # 输出重复项到txt文件
            txt_file_path = f"{os.path.splitext(file_path)[0]}.txt"
            with open(txt_file_path, 'w') as txt_file:
                for value, count in duplicates.items():
                    txt_file.write(f"{value}: {count}\n")

            # 清空原始数据
            sheet.delete_rows(1, sheet.max_row)
            # 写入处理后的数据
            for row in rows_to_keep.values():
                sheet.append(row)

    # 输出文件到原文件目录下，并加上change标记
    file_name, file_ext = os.path.splitext(file_path)
    output_file_path = f"{file_name}_change{file_ext}"
    wb.save(output_file_path)
    print(f"处理完成：{output_file_path}")
    print(f"重复项已写入：{txt_file_path}")


def main():
    directory = r"D:\test4"
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            process_excel_file(file_path)


if __name__ == "__main__":
    main()
