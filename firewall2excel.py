import os
from openpyxl import load_workbook


def replace_values_in_column(ws, column, replacement_dict):
    for cell in ws[column]:
        if cell.value in replacement_dict:
            cell.value = replacement_dict[cell.value]


def main():
    # 文件路径
    directory = "/Users/kuihe/Python/test4/"
    file_name = "CNSZ02-3E05-13U-RH-SRX4100-M 测试.xlsx"
    file_path = os.path.join(directory, file_name)

    # 加载工作簿
    wb = load_workbook(file_path)

    # 获取“策略”工作表
    ws_strategy = wb["策略"]

    # 获取“地址组对象”工作表
    ws_address = wb["地址组对象"]

    # 获取“地址对象”工作表
    ws_ipaddress = wb["地址对象"]

    # 获取“服务组对象”工作表
    ws_service = wb["服务组对象"]

    # 将工作表“策略”中D列的数据复制到E列
    for row in range(2, ws_strategy.max_row + 1):
        ws_strategy[f"E{row}"].value = ws_strategy[f"D{row}"].value

    # 将工作表“策略”中G列的数据复制到H列
    for row in range(2, ws_strategy.max_row + 1):
        ws_strategy[f"H{row}"].value = ws_strategy[f"G{row}"].value

    # 构建替换字典
    replacement_dict = {ws_address[f"A{i}"].value: ws_address[f"B{i}"].value for i in range(2, 35001)}

    # 替换多行值
    for row in range(1, 30001):
        cell_value = ws_strategy[f"E{row}"].value
        if isinstance(cell_value, str) and "\n" in cell_value:
            lines = cell_value.split("\n")
            for i, line in enumerate(lines):
                if line in replacement_dict:
                    lines[i] = replacement_dict[line]
                else:
                    # 如果没有匹配到，保持值不变
                    lines[i] = line
            ws_strategy[f"E{row}"].value = "\n".join(lines)

    # 替换单行值
    replace_values_in_column(ws_strategy, "E", replacement_dict)

    # 处理工作表“策略”中H列的数据
    for row in range(1, 30001):
        cell_value = ws_strategy[f"H{row}"].value
        if isinstance(cell_value, str) and "\n" in cell_value:
            lines = cell_value.split("\n")
            for i, line in enumerate(lines):
                if line in replacement_dict:
                    lines[i] = replacement_dict[line]
                else:
                    # 如果没有匹配到，保持值不变
                    lines[i] = line
            ws_strategy[f"H{row}"].value = "\n".join(lines)

    # 替换单行值
    replace_values_in_column(ws_strategy, "H", replacement_dict)





    # 将工作表“策略”中E列的数据复制到F列
    for row in range(2, ws_strategy.max_row + 1):
        ws_strategy[f"F{row}"].value = ws_strategy[f"E{row}"].value

    # 将工作表“策略”中H列的数据复制到I列
    for row in range(2, ws_strategy.max_row + 1):
        ws_strategy[f"I{row}"].value = ws_strategy[f"H{row}"].value

    # 构建替换字典
    replacementip_dict = {ws_ipaddress[f"A{i}"].value: ws_ipaddress[f"B{i}"].value for i in range(2, 35001)}

    # 替换多行值
    for row in range(1, 30001):
        cell_value = ws_strategy[f"F{row}"].value
        if isinstance(cell_value, str) and "\n" in cell_value:
            lines = cell_value.split("\n")
            for i, line in enumerate(lines):
                if line in replacementip_dict:
                    lines[i] = replacementip_dict[line]
            ws_strategy[f"F{row}"].value = "\n".join(lines)

    # 替换单行值
    replace_values_in_column(ws_strategy, "F", replacementip_dict)

    # 替换多行值
    for row in range(1, 30001):
        cell_value = ws_strategy[f"I{row}"].value
        if isinstance(cell_value, str) and "\n" in cell_value:
            lines = cell_value.split("\n")
            for i, line in enumerate(lines):
                if line in replacementip_dict:
                    lines[i] = replacementip_dict[line]
            ws_strategy[f"I{row}"].value = "\n".join(lines)

    # 替换单行值
    replace_values_in_column(ws_strategy, "I", replacementip_dict)





    # 将工作表“策略”中D列的数据复制到E列
    for row in range(2, ws_strategy.max_row + 1):
        ws_strategy[f"K{row}"].value = ws_strategy[f"J{row}"].value


    # 构建替换字典
    replacement_service_dict = {ws_service[f"A{i}"].value: ws_service[f"B{i}"].value for i in range(2, 35001)}


    # 替换多行值
    for row in range(1, 30001):
        cell_value = ws_strategy[f"K{row}"].value
        if isinstance(cell_value, str) and "\n" in cell_value:
            lines = cell_value.split("\n")
            for i, line in enumerate(lines):
                if line in replacement_service_dict:
                    lines[i] = replacement_service_dict[line]
            ws_strategy[f"K{row}"].value = "\n".join(lines)

    # 替换单行值
    replace_values_in_column(ws_strategy, "K", replacement_service_dict)




    # 保存更改
    changed_file_name = os.path.splitext(file_name)[0] + "_change.xlsx"
    changed_file_path = os.path.join(directory, changed_file_name)
    wb.save(changed_file_path)
    print(f"Excel文件已保存为：{changed_file_path}")


if __name__ == "__main__":
    main()
